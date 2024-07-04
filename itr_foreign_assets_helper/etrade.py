import datetime
import logging
import typing

import openpyxl
import openpyxl.worksheet
import openpyxl.worksheet.worksheet

from pathlib import Path

from . import forex
from . import stock
from . import utils


logger = logging.getLogger(__name__)


class ETradeTransactions:
    
    def __init__(self, holdings_file: typing.IO, sale_transactions_file: typing.IO, sbi_reference_rates: forex.SBIReferenceRates) -> None:
        self.broker = "ETrade"
        self.sbi_reference_rates = sbi_reference_rates
        self.shares_issued = self.__get_shares_issued(holdings_file=holdings_file)
        self.shares_sold = self.__get_shares_sold(sale_transactions_file=sale_transactions_file)
        self.cash = self.__get_cash(holdings_file=holdings_file)
    
    def __extract_data(
            self,
            file: typing.IO,
            sheet_name: str,
            row_to_skip_at_start: int,
            row_to_skip_at_end: int,
            required_columns: typing.Dict[str, str]
    ) -> typing.Dict:
        file_name = Path(file.name).name
        wb = openpyxl.load_workbook(file)
        ws = wb[sheet_name]
        required_column_letters = utils.get_ws_column_metadata(
            ws=ws,
            columns=required_columns.keys(),
        )
        data = []
        # mix_row and max_row are 1-based index.
        min_row = ws.min_row + row_to_skip_at_start
        max_row = ws.max_row - row_to_skip_at_end
        curr_row_num = min_row
        for row in ws.iter_rows(min_row=min_row, max_row=max_row):
            curr_row_data = {}
            curr_row_data['source_metadata'] = {
                'file_name': file_name,
                'sheet_name': sheet_name,
                'row': curr_row_num,
            }
            for cell in row:
                if cell.column_letter in required_column_letters:
                    cell_title = required_column_letters[cell.column_letter]
                    translated_title = required_columns[cell_title]
                    curr_row_data[translated_title] = cell.value
            data.append(curr_row_data)
            curr_row_num += 1
        return data
    
    def __get_shares_issued(self, holdings_file: typing.IO) -> typing.List[stock.ShareIssuedRecord]:
        sheet_name = 'Sellable'
        required_columns = {
            'Symbol': 'ticker',
            'Sellable Qty.': 'shares_issued',
            'Grant Number': 'award_number',
            'Release Date': 'issue_date',
            'Purchase Date FMV': 'fmv_per_share_on_issue_date',
        }
        share_issued_records = []
        # the first row is the column names and hence skipped
        row_to_skip_at_start = 1
        # the last row is the total and hence skipped
        row_to_skip_at_end = 1
        raw_stocks_data = self.__extract_data(
            file=holdings_file,
            sheet_name=sheet_name,
            required_columns=required_columns,
            row_to_skip_at_start=row_to_skip_at_start,
            row_to_skip_at_end=row_to_skip_at_end,
        )
        for stock_data in raw_stocks_data:
            share_issued_record = stock.ShareIssuedRecord(
                sbi_reference_rates=self.sbi_reference_rates,
                source_metadata=stock_data['source_metadata'],
                broker=self.broker,
                comments='',
                ticker=stock_data['ticker'],
                award_number=int(stock_data['award_number']),
                shares_issued=stock_data['shares_issued'],
                issue_date=datetime.datetime.strptime(stock_data['issue_date'], '%d-%b-%Y').date(),
                fmv_per_share_on_issue_date=float(str(stock_data['fmv_per_share_on_issue_date']).replace('$', '')),
            )
            logger.debug(share_issued_record)
            share_issued_records.append(share_issued_record)
        return share_issued_records

    def __get_shares_sold(self, sale_transactions_file: typing.IO) -> typing.List[stock.ShareSoldRecord]:
        sheet_name = 'G&L_Expanded'
        required_columns = {
            'Symbol': 'ticker',
            'Qty.': 'shares_sold',
            'Grant Number': 'award_number',
            'Date Acquired': 'issue_date',
            'Vest Date FMV': 'fmv_per_share_on_issue_date',
            'Date Sold': 'sale_date',
            'Proceeds Per Share': 'fmv_per_share_on_sale_date',
            'Order Number': 'sale_order_number'
        }
        share_sold_records = []
        # the first row is the column names and hence skipped
        # the second row is a summary rpw and hence skipped
        row_to_skip_at_start = 2
        # there is no extra row at the end. hence no rows skipped.
        row_to_skip_at_end = 0
        raw_stocks_data = self.__extract_data(
            file=sale_transactions_file,
            sheet_name=sheet_name,
            required_columns=required_columns,
            row_to_skip_at_start=row_to_skip_at_start,
            row_to_skip_at_end=row_to_skip_at_end,
        )
        for stock_data in raw_stocks_data:
            share_sold_record = stock.ShareSoldRecord(
                sbi_reference_rates=self.sbi_reference_rates,
                source_metadata=stock_data['source_metadata'],
                broker=self.broker,
                comments='',
                ticker=stock_data['ticker'],
                award_number=int(stock_data['award_number']),
                issue_date=datetime.datetime.strptime(stock_data['issue_date'], '%m/%d/%Y').date(),
                fmv_per_share_on_issue_date=float(str(stock_data['fmv_per_share_on_issue_date']).replace('$', '')),
                shares_sold=stock_data['shares_sold'],
                sale_date=datetime.datetime.strptime(stock_data['sale_date'], '%m/%d/%Y').date(),
                fmv_per_share_on_sale_date=float(str(stock_data['fmv_per_share_on_sale_date']).replace('$', '')),
            )
            logger.debug(share_sold_record)
            share_sold_records.append(share_sold_record)
        return share_sold_records
    
    def __get_cash(self, holdings_file: typing.IO) -> stock.CashRecord:
        sheet_name = 'Other Holdings'
        required_columns = {
            'Est. Market Value': 'amount',
        }
        cash = None
        # the first row is the column names and hence skipped
        # the second row is empty with lines and hence skipped
        row_to_skip_at_start = 2
        # the last row caontains the cash as total
        row_to_skip_at_end = 0
        raw_cash_data = self.__extract_data(
            file=holdings_file,
            sheet_name=sheet_name,
            required_columns=required_columns,
            row_to_skip_at_start=row_to_skip_at_start,
            row_to_skip_at_end=row_to_skip_at_end,
        )
        if len(raw_cash_data) != 1:
            raise Exception(f'Failed to extract cash holdings for {self.broker}')
        cash = stock.CashRecord(
            sbi_reference_rates=self.sbi_reference_rates,
            source_metadata=raw_cash_data[0]['source_metadata'],
            broker=self.broker,
            comments='',
            amount=float(str(raw_cash_data[0]['amount']).replace('$', '')),
        )
        logger.debug(cash)
        return cash
