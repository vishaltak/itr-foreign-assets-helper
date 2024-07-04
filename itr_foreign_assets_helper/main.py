import argparse
import datetime
import logging
import re
import typing

import openpyxl

from pathlib import Path

from . import etrade
from . import forex
from . import itr_schedule_fa


logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(".log.txt"),
        logging.StreamHandler()
    ]
)

if logging.root.level == logging.DEBUG:
    # following modules create a lot of noise on debug. disable them.
    logging.getLogger('yfinance').setLevel(logging.WARNING)
    logging.getLogger('peewee').setLevel(logging.WARNING)
    logging.getLogger('urllib3').setLevel(logging.WARNING)

logger = logging.getLogger(__name__)

def validate_financial_year_input(input) -> typing.Tuple[datetime.date, datetime.date]:
    match = re.match(r'^(\d{4})-(\d{4})$', input)
    if not match:
        raise argparse.ArgumentTypeError(f'Invalid financial year format: {input}. Expected format is YYYY-YYYY.')
    start_year, end_year = map(int, match.groups())
    if end_year != start_year + 1:
        raise argparse.ArgumentTypeError(f'Invalid financial year format: {input}. Gap between years should only be 1.')
    financial_year_start_date = datetime.date(year=start_year, month=4, day=1)
    financial_year_end_year = datetime.date(year=end_year, month=3, day=31)
    return financial_year_start_date, financial_year_end_year

def main():
    parser = argparse.ArgumentParser(
        prog='itr-foreign-assets-helper',
        description='itr-foreign-assets-helper',
    )
    parser.add_argument('--financial-year',
        type=validate_financial_year_input,
        required=True,
        help='Financial year for which to generate data. e.g. 2023-2024'    
    )
    parser.add_argument('--sbi-reference-rates',
        type=argparse.FileType('r'),
        required=False,
        help='SBI Reference Rates CSV. If not specified, download from https://github.com/sahilgupta/sbi-fx-ratekeeper/blob/main/csv_files/SBI_REFERENCE_RATES_USD.csv'
    )
    parser.add_argument('--etrade-holdings',
        type=argparse.FileType('rb'),
        required=True,
        help='ETrade holdings file at the end of financial year i.e. right after 31 March'
    )
    parser.add_argument('--etrade-sale-transactions',
        type=argparse.FileType('rb'),
        required=True,
        help='ETrade sale transactions file for financial year i.e. 1 January to 31 March(next year). The records for 1 January to 31 March(end of previous Indian financial year year) are required for Schedule FA and the records from 1 April(start of current Indian financial year year) to 31 March(end of Indian financial year year) are required for Schedule CG and Schedule AL.'
    )
    args = parser.parse_args()

    sbi_reference_rates = forex.SBIReferenceRates(args.sbi_reference_rates)

    etrade_transcations = etrade.ETradeTransactions(
        holdings_file=args.etrade_holdings,
        sale_transactions_file=args.etrade_sale_transactions,
        sbi_reference_rates=sbi_reference_rates,
    )
    logger.debug(etrade_transcations.shares_issued)
    logger.debug(etrade_transcations.shares_sold)

    itr_schedule_fa_a3 = itr_schedule_fa.ScheduleFAA3(
        shares_issued=etrade_transcations.shares_issued,
        shares_sold=etrade_transcations.shares_sold,
        sbi_reference_rates=sbi_reference_rates,
        financial_year=args.financial_year
    )

    itr_schedule_fa_a2 = itr_schedule_fa.ScheduleFAA2(
        cash_record=etrade_transcations.cash,
        sbi_reference_rates=sbi_reference_rates,
        financial_year=args.financial_year
    )

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    logger.info('ITR Data')

    logger.info('Schedule FA A3')
    logger.info(itr_schedule_fa_a3.entries)
    sheet_name = 'Schedule FA A3'
    workbook.create_sheet(title=sheet_name)
    itr_schedule_fa_a3.export(workbook=workbook, sheet_name=sheet_name)
    
    logger.info('Schedule FA A2')
    logger.info(itr_schedule_fa_a2.entries)
    sheet_name = 'Schedule FA A2'
    workbook.create_sheet(title=sheet_name)
    itr_schedule_fa_a2.export(workbook=workbook, sheet_name=sheet_name)

    file_name = f'itr-helper-fy-{args.financial_year[0].year}-{args.financial_year[1].year}.xlsx'
    file_path = Path(__file__).resolve().parent.parent / 'output' / file_name
    file_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(file_path)
    workbook.close()
